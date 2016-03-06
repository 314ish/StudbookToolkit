from openpyxl import Workbook
from openpyxl import load_workbook


class excelWriter:

    def __init__(self, name):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.name = name

    # studbook must be studbook type (duh)
    def writeStudbook(self, studbook):
        self.writeMultipleRows(studbook.header)

        for record in studbook.directory:
            self.writeMultipleRows(record.returnExcelFormat())

    # input data will be written one element per row (so use a list of lists for 2D writing)
    def writeMultipleRows(self, data):
        for row in data:
            self.writeRow(row)

    # data must be a simple list
    def writeRow(self, data):
        self.ws.append(data)

    def close(self):
        self.wb.save(self.name+".xlsx")


class excelReader:

    def __init__(self, name):
        self.wb = load_workbook(filename = name)
        self.sheet_ranges = self.wb[self.wb.get_sheet_names()[0]]

    # yeah, this only gets B & C columns from the sheet, you've been warned
    def getRecordsAsList(self):
        returnMe = []
        i = 2
        while i <= len(self.sheet_ranges.rows):
            sire = self.sheet_ranges['B'+str(i)].value
            dam = self.sheet_ranges['C'+str(i)].value
            sd = [sire, dam]
            returnMe.append(sd)
            i += 1

        return returnMe
