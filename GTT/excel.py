from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelWriter:
    """These objects can be used to write Excel formatted files. Please note
    that all data is stored in RAM until close: is called which makes this
    potentially 'difficult' to use with very large data sets.
    """

    def __init__(self, name):
        """Initialize an ExcelWriter: object.

        Args:
           name (str):  name of the file to create when writing (include .xlsx
           in the variable if you want correct Excel naming!)

        Returns:
           ExcelWriter:
        """

        self.wb = Workbook()
        self.ws = self.wb.active
        self.name = name

    def write_studbook(self, studbook):
        """write the entire contents from a Studbook: into this object.

        Args:
           studbook (Studbook):  Studbook: to be written to this object

        Returns:
            nothing
        """

        self.ws.append(studbook.header)

        for record in studbook.directory:
            self._write_row(record.get_record(), studbook.header)

    def _write_multiple_rows(self, data):
        """append multiple rows into this object from a python list

        Args:
           data (list):  each element of data: must contain a list which represents a row of data

        Returns:
           nothing
        """

        for row in data:
            self._write_row(row)

    # data must be a simple list
    def _write_row(self, data, fields):
        """add a single studbookRecord into this object using the fields
         argument to identify which data should be included

        Args:
           data (dict):  a dictionary with all possible elements
           fields (list): a list of keys in the data dictionary to write (in-order)

        Returns:
            nothing
        """
        write_me = []
        for key in fields:
            write_me.append(data[key])

        self.ws.append(write_me)

    def close(self):
        """save all the data from this object into an Excel file. This may be
        poorly named, but we don't really have a close-out function yet other
        than this.

        Returns:
            nothing
        """

        self.wb.save(self.name)


class ExcelReader:
    """These objects can be used to read Excel formatted files.
    """

    def __init__(self, name):
        """Initialize an ExcelReader: object.

        Args:
           name (str):  name of the file to read from

        Returns:
           ExcelWriter:
        """
        self.wb = load_workbook(filename=name)
        self.sheet_ranges = self.wb[self.wb.get_sheet_names()[0]]

    def get_records_as_list(self):
        """read all records from this file into a (2-dimensional) list. Each
        element of the list is list representing an entire row of data. In
        each internal list each element is a single column.

        FOR NOW THIS ONLY READS THE 'B' AND 'C' COLUMNS

        Returns:
           returnMe (list):
        """

        return_me = []
        i = 2
        while i <= len(self.sheet_ranges.rows):
            sire = self.sheet_ranges['B'+str(i)].value
            dam = self.sheet_ranges['C'+str(i)].value
            sd = [sire, dam]
            return_me.append(sd)
            i += 1

        return return_me

